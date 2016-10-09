#!/usr/bin/env python

import argparse
import json
import os
import pickle
import sys

import openpyxl

def get_ensemble_dict(ensemble_gff3):
    pickle_file = ensemble_gff3 + '.pickle'
    if os.path.exists(pickle_file):
        with open(pickle_file, 'rb') as f_open:
            ensemble_dict = pickle.load(f_open)
            return ensemble_dict
    ensemble_dict = dict()
    with open(ensemble_gff3, 'r') as f_open:
        for line in f_open:
            if line.startswith('#'):
                continue
            line_split = line.split('\t')
            data = line_split[8]
            data_split = data.split(';')
            gene_id = str()
            gene_name = str()
            for keyvalue in data_split:
                keyvalue_split = keyvalue.split('=')
                key = keyvalue_split[0]
                value = keyvalue_split[1]
                if key == 'gene_id':
                    gene_id = value
                if key == 'Name':
                    gene_name = value
            ensemble_dict[gene_id] = gene_name
    with open(pickle_file, 'wb') as f_open:
        pickle.dump(ensemble_dict, f_open)
    return ensemble_dict

def get_gencode_dict(gencode_gff3):
    pickle_file = gencode_gff3 + '.pickle'
    if os.path.exists(pickle_file):
        with open(pickle_file, 'rb') as f_open:
            idname_dict = pickle.load(f_open)
            return idname_dict
    gencode_dict = dict()
    with open(gencode_gff3, 'r') as f_open:
        for line in f_open:
            if line.startswith('#'):
                continue
            line_split = line.split('\t')
            data = line_split[8]
            data_split = data.split(';')
            gene_id = str()
            gene_name = str()
            for keyvalue in data_split:
                keyvalue_split = keyvalue.split('=')
                key = keyvalue_split[0]
                value = keyvalue_split[1]
                if key == 'ID':
                    gene_id = value
                if key == 'gene_name':
                    gene_name = value
            gencode_dict[gene_id] = gene_name
    with open(pickle_file, 'wb') as f_open:
        pickle.dump(gencode_dict, f_open)
    return gencode_dict

def get_hgnc_dict(hgnc_json):
    hgnc_dict = dict()
    with open(hgnc_json) as f_open:
        data = json.load(f_open)
    for item in data['response']['docs']:
        try:
            ensembl_gene_id = item['ensembl_gene_id']
            symbol = item['symbol']
            hgnc_dict[ensembl_gene_id] = symbol
        except:
            print('item: %s' % item)
            continue
    return hgnc_dict

def write_names(xlsx_file, ensemble_dict, gencode_dict, hgnc_dict):
    wb_in = openpyxl.load_workbook(filename=xlsx_file, read_only=True)
    ws_in = wb_in.active
    wb_basename, ext = os.path.splitext(os.path.basename(xlsx_file))
    wb_out_name = wb_basename + '_genename.xlsx'
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet()

    missing_name_list = list()
    ws_out.append(['gene_id', 'gene_name'])
    for idx, row in enumerate(ws_in.rows, 1):
        for cell in row:
            if cell.value.startswith('ENSG'):
                if cell.value in ensemble_dict:
                    ws_out.append([cell.value, ensemble_dict[cell.value]])
                elif cell.value in hgnc_dict:
                    ws_out.append([cell.value, hgnc_dict[cell.value]])
                elif cell.value in gencode_dict:
                    ws_out.append([cell.value, gencode_dict[cell.value]])
                else:
                    missing_name_list.append(cell.value)
    wb_out.save(wb_out_name)

    print('len(missing_name_list)=%s' % len(missing_name_list))
    wb_missing = openpyxl.Workbook(write_only=True)
    ws_missing = wb_missing.create_sheet()
    ws_missing.append(['gene_id'])
    for missing_name in missing_name_list:
        ws_missing.append([missing_name])
    wb_missing.save('missing_genename.xlsx')
    return

def main():
    parser = argparse.ArgumentParser('convert readgroups to json')

    parser.add_argument('--ensemble_gff3',
                        required = True
    )
    parser.add_argument('--gencode_gff3',
                        required = True,
    )
    parser.add_argument('--hgnc_json',
                        required = True,
    )
    parser.add_argument('-x', '--xlsx_file',
                        required = True,
                        help = 'BAM file.'
    )

    args = parser.parse_args()
    ensemble_gff3 = args.ensemble_gff3
    gencode_gff3 = args.gencode_gff3
    hgnc_json = args.hgnc_json
    xlsx_file = args.xlsx_file

    ensemble_dict = get_ensemble_dict(ensemble_gff3)
    gencode_dict = get_gencode_dict(gencode_gff3)
    hgnc_dict = get_hgnc_dict(hgnc_json)
    write_names(xlsx_file, ensemble_dict, gencode_dict, hgnc_dict)
    return

if __name__ == '__main__':
    main()
